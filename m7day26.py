# ==========================================================
# Month 7 - Day 26
# CSV, JSON & Excel File Processing
#
# Topics Covered:
# 1. Write CSV File
# 2. Read CSV File
# 3. DictReader & DictWriter
# 4. JSON File Processing
# 5. Create Excel File
# 6. Read Excel File
# ==========================================================

import csv
import json


# ==========================================================
# 1. WRITE CSV FILE
# ==========================================================

print("=" * 60)
print("1. WRITE CSV FILE")
print("=" * 60)

students = [
    ["ID", "Name", "Branch", "CGPA"],
    [1, "Arjun", "CSE", 9.1],
    [2, "Rahul", "ISE", 8.7],
    [3, "Priya", "ECE", 9.3],
    [4, "Anjali", "AIML", 8.9]
]

with open(
    "students.csv",
    "w",
    newline="",
    encoding="utf-8"
) as file:

    writer = csv.writer(file)

    writer.writerows(students)

print("students.csv created successfully!")


# ==========================================================
# 2. READ CSV FILE
# ==========================================================

print("\n" + "=" * 60)
print("2. READ CSV FILE")
print("=" * 60)

with open(
    "students.csv",
    "r",
    encoding="utf-8"
) as file:

    reader = csv.reader(file)

    for row in reader:
        print(row)


# ==========================================================
# 3. DICTREADER & DICTWRITER
# ==========================================================

print("\n" + "=" * 60)
print("3. DICTREADER & DICTWRITER")
print("=" * 60)

employees = [
    {
        "id": 101,
        "name": "Amit",
        "department": "Development"
    },
    {
        "id": 102,
        "name": "Neha",
        "department": "Testing"
    },
    {
        "id": 103,
        "name": "Kiran",
        "department": "Data Science"
    }
]

fieldnames = [
    "id",
    "name",
    "department"
]

with open(
    "employees.csv",
    "w",
    newline="",
    encoding="utf-8"
) as file:

    writer = csv.DictWriter(
        file,
        fieldnames=fieldnames
    )

    writer.writeheader()

    writer.writerows(employees)

print("employees.csv created successfully!")


print("\nEmployee Records:")

with open(
    "employees.csv",
    "r",
    encoding="utf-8"
) as file:

    reader = csv.DictReader(file)

    for row in reader:
        print(row)


# ==========================================================
# 4. JSON FILE PROCESSING
# ==========================================================

print("\n" + "=" * 60)
print("4. JSON FILE PROCESSING")
print("=" * 60)

course_data = {
    "course": "Python Mastery",
    "month": 7,
    "topics": [
        "Python",
        "DSA",
        "APIs",
        "Database",
        "File Processing"
    ]
}


# ----------------------------------------------------------
# WRITE JSON FILE
# ----------------------------------------------------------

with open(
    "course.json",
    "w",
    encoding="utf-8"
) as file:

    json.dump(
        course_data,
        file,
        indent=4
    )

print("course.json created successfully!")


# ----------------------------------------------------------
# READ JSON FILE
# ----------------------------------------------------------

with open(
    "course.json",
    "r",
    encoding="utf-8"
) as file:

    loaded_data = json.load(file)

print("\nJSON Data:")

print("Course :", loaded_data["course"])
print("Month  :", loaded_data["month"])

print("Topics:")

for topic in loaded_data["topics"]:
    print("-", topic)


# ==========================================================
# 5. EXCEL FILE CREATION
# ==========================================================

print("\n" + "=" * 60)
print("5. CREATE EXCEL FILE")
print("=" * 60)

try:

    from openpyxl import Workbook
    from openpyxl import load_workbook

    workbook = Workbook()

    # workbook.active may be Optional according to type checking.
    # Instead, create our own worksheet explicitly.

    default_sheet = workbook.active

    if default_sheet is not None:
        workbook.remove(default_sheet)

    sheet = workbook.create_sheet(
        title="Students"
    )

    # Add heading

    sheet.append([
        "ID",
        "Name",
        "Branch",
        "CGPA"
    ])

    # Add student records

    sheet.append([
        1,
        "Arjun",
        "CSE",
        9.1
    ])

    sheet.append([
        2,
        "Rahul",
        "ISE",
        8.7
    ])

    sheet.append([
        3,
        "Priya",
        "ECE",
        9.3
    ])

    sheet.append([
        4,
        "Anjali",
        "AIML",
        8.9
    ])

    # Save Excel file

    workbook.save(
        "students.xlsx"
    )

    print(
        "students.xlsx created successfully!"
    )


    # ======================================================
    # 6. READ EXCEL FILE
    # ======================================================

    print("\n" + "=" * 60)
    print("6. READ EXCEL FILE")
    print("=" * 60)

    workbook = load_workbook(
        "students.xlsx"
    )

    sheet = workbook["Students"]

    print("Excel Student Records:\n")

    for row in sheet.iter_rows(
        values_only=True
    ):
        print(row)

    workbook.close()


except ImportError:

    print("""
openpyxl is not installed.

Install it using:

python -m pip install openpyxl

Then run:

python m7day26.py
""")


except PermissionError:

    print("""
Permission Error!

students.xlsx may already be
open in Microsoft Excel.

Close the Excel file and
run the program again.
""")


except Exception as error:

    print(
        "Excel Processing Error:",
        error
    )


# ==========================================================
# 7. CSV DATA ANALYSIS
# ==========================================================

print("\n" + "=" * 60)
print("7. CSV DATA ANALYSIS")
print("=" * 60)

total_cgpa = 0
student_count = 0
highest_cgpa = 0
top_student = ""

with open(
    "students.csv",
    "r",
    encoding="utf-8"
) as file:

    reader = csv.DictReader(file)

    for row in reader:

        cgpa = float(row["CGPA"])

        total_cgpa += cgpa
        student_count += 1

        if cgpa > highest_cgpa:

            highest_cgpa = cgpa

            top_student = row["Name"]


if student_count > 0:

    average_cgpa = (
        total_cgpa / student_count
    )

    print(
        "Total Students :",
        student_count
    )

    print(
        "Average CGPA   :",
        round(average_cgpa, 2)
    )

    print(
        "Highest CGPA   :",
        highest_cgpa
    )

    print(
        "Top Student    :",
        top_student
    )


# ==========================================================
# 8. JSON STRING CONVERSION
# ==========================================================

print("\n" + "=" * 60)
print("8. JSON STRING CONVERSION")
print("=" * 60)

student = {
    "id": 1,
    "name": "Arjun",
    "branch": "CSE",
    "cgpa": 9.1
}


# Python Dictionary -> JSON String

json_string = json.dumps(
    student,
    indent=4
)

print("Python Dictionary -> JSON String")

print(json_string)


# JSON String -> Python Dictionary

python_dictionary = json.loads(
    json_string
)

print("\nJSON String -> Python Dictionary")

print(python_dictionary)


# ==========================================================
# INTERVIEW SUMMARY
# ==========================================================

print("\n" + "=" * 60)
print("INTERVIEW SUMMARY")
print("=" * 60)

print("""
✔ CSV

CSV =
Comma-Separated Values

Python provides the built-in:

csv module

--------------------------------------------------

✔ csv.writer()

Used to write CSV data.

writer.writerow()

Writes one row.

writer.writerows()

Writes multiple rows.

--------------------------------------------------

✔ csv.reader()

Reads CSV rows as lists.

--------------------------------------------------

✔ DictWriter

Writes dictionaries
to CSV files.

Example:

csv.DictWriter()

--------------------------------------------------

✔ DictReader

Reads each CSV row
as a dictionary.

This makes accessing columns
by name easier.

Example:

row["Name"]

--------------------------------------------------

✔ JSON

JSON =
JavaScript Object Notation

Commonly used for:

• APIs
• Configuration
• Data Exchange
• Web Applications

--------------------------------------------------

✔ json.dump()

Python Object
        ↓
JSON File

--------------------------------------------------

✔ json.dumps()

Python Object
        ↓
JSON String

--------------------------------------------------

✔ json.load()

JSON File
        ↓
Python Object

--------------------------------------------------

✔ json.loads()

JSON String
        ↓
Python Object

--------------------------------------------------

✔ Excel

Python Library:

openpyxl

Install using:

python -m pip install openpyxl

--------------------------------------------------

✔ Workbook

Create Excel workbook:

Workbook()

--------------------------------------------------

✔ Worksheet

Create worksheet:

workbook.create_sheet()

--------------------------------------------------

✔ Append Rows

sheet.append()

--------------------------------------------------

✔ Save Excel

workbook.save(
    "students.xlsx"
)

--------------------------------------------------

✔ Read Excel

load_workbook(
    "students.xlsx"
)

--------------------------------------------------

✔ iter_rows()

Used to iterate through
Excel rows.

values_only=True

returns cell values directly.

--------------------------------------------------

CSV vs JSON vs Excel

CSV
→ Simple tabular data

JSON
→ Structured / nested data

Excel
→ Reports and spreadsheets

SQLite
→ Database storage

--------------------------------------------------

Most Asked Interview Questions

✔ CSV vs JSON

✔ CSV vs Excel

✔ reader() vs DictReader()

✔ writer() vs DictWriter()

✔ dump() vs dumps()

✔ load() vs loads()

✔ File Encoding

✔ Context Managers

✔ openpyxl

✔ Reading large files
""")


print("\n" + "=" * 60)
print("DAY 26 COMPLETED")
print("=" * 60)

print("""
Files Created:

1. students.csv
2. employees.csv
3. course.json
4. students.xlsx

Concepts Practiced:

✔ CSV
✔ JSON
✔ Excel
✔ File Handling
✔ Data Processing
✔ Error Handling
""")